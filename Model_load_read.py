"""读取文件"""
from GeneralDesign import SIMS
import os
from openpyxl import load_workbook
class Read(SIMS):

    def __init__(self,file_name):
        self.file_name = './学生数据文件/' + file_name
        self.read_cache = []
    def read_data_to_cache(self):
        if ".txt" in self.file_name:
            print("读取txt文件完成!")
            self._read_data_by_txt()
            return self.read_cache
        else:
            print("读取excel文件完成!")
            self._read_data_by_excel()
            return self.read_cache

    def _read_data_by_txt(self):
        with open(self.file_name,'r',encoding="utf-8") as f:
            for line in f.readlines():
                line = line[0] + line[3:].strip('\n')
                self.read_cache.append(line)
            return self.read_cache

    def _read_data_by_excel(self):
        wb = load_workbook("./学生数据文件/学生数据文件.xlsx")
        ws = wb.active
        for row in ws.iter_rows(values_only=True):
            self.read_cache.append(row)
        # print(self.read_cache)

# 加载文件数据
def loading():
    file_txt = "学生数据文件.txt"
    file_xlsx = "学生数据文件.xlsx"

    if file_txt in os.listdir("./学生数据文件/"):
        read = Read(file_txt)
        loading_data_list = read.read_data_to_cache()
        return loading_data_list
    elif file_xlsx in os.listdir("./学生数据文件/"):
        read = Read(file_xlsx)
        loading_data_list = read.read_data_to_cache()
        return loading_data_list
    else:
        return None

def read_max_10_step():
    loading_data_list = loading()
    try:
        if loading_data_list == None:
            print("暂无数据")
        else:
            step = 10
            index = 0
            while index < len(loading_data_list):
                for i in range(index, min(index + step, len(loading_data_list))):
                    show_better(loading_data_list[i])
                index += step
                if index < len(loading_data_list):
                    go_on_read = input("是否继续读取?[y/n]")
                    if go_on_read != 'y':
                        break
    except:
        print("请检查文件是否存在")

# 后续添加显示更人性化函数 pass to do it 
def show_better(data):
    print(data)
