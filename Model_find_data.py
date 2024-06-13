"""找到我想要的数据"""
import os
from openpyxl import load_workbook
from GeneralDesign import SIMS
from Model_menu import Menu
from Model_load_read import loading
class Find_or_Drop(SIMS):

    def __init__(self):
        self.file_name_txt = "学生数据文件.txt"
        self.find_name = None
        self.find_id = None
        self.loading_data_list = loading()

    def get_find_id(self,sid):
        self.find_id = sid
        return self.find_id

    def get_find_name(self,name):
        self.find_name = name
        return self.find_name

    def search_information(self):
        while True:
            _ = Menu().menu_search()
            find_op = input("请选择你的操作:")
            if find_op == '3':
                break
            self.knew_txt_or_excel()

            if self.model == None:
                print("暂无数据")
                break
            if self.model == 'txt':
                if find_op == '1':
                    sid = input("请输入你要查询的学号:")
                    self.find_id = self.get_find_id(sid)
                    self.find_by_id()  # 调用find_by_id方法
                    continue
                elif find_op == '2':
                    sname = input("请输入你要查询的学生姓名:")
                    self.find_name = self.get_find_name(sname)
                    self.find_by_name()
                    continue
            elif self.model == 'excel':
                if find_op == '1':
                    sid = input("请输入你要查询的学号:")
                    self.find_id = self.get_find_id(sid)
                    self.find_by_id()  # 调用find_by_id方法
                    continue
                elif find_op == '2':
                    sname = input("请输入你要查询的学生姓名:")
                    self.find_name = self.get_find_name(sname)
                    self.find_by_name()
                    continue
            else:
                print("over")
                break

    # 用与处理excel文件
    def tuple_to_dict(self,data):
        keys = data[0]
        result = [dict(zip(keys, values)) for values in data[1:]]
        return result

    # 用与处理txt文件
    def list_to_dict(self,data):
        # 找到"学号"的位置
        start_index = data.find("学号")
        # 从"学号"开始切割数据
        data = data[start_index:]
        # 将数据转换成字典形式
        items = data.split(",,'")
        result = {}
        for item in items:
            key_value = item.split(":,'")
            if len(key_value) == 2:
                key = key_value[0].strip("'")
                value = key_value[1].strip("'")
                result[key] = value
        return result

    """
    添加行参model 用于识别采用哪种方法 txt | excel
    """
    # 后续设计：处理txt或excel单独存在时，使用方法冲突
    # 用 os 去识别文件
    def knew_txt_or_excel(self):
        files = os.listdir("./学生数据文件/")
        if len(files) != 0:
            for file in files:
                if file.endswith('.txt'):
                    self.model = 'txt'
                    return self.model
                elif file.endswith('.xlsx'):
                    self.model = 'excel'
                    return self.model
        else:
            self.model = None
    def find_by_name(self):
        found = False  # 添加一个标记来表示是否找到
        if self.model == 'txt':
            for _ in self.loading_data_list:
                dict_data = self.list_to_dict(_)
                if self.find_name == dict_data["姓名"]:
                    print("成功查询")
                    found = True
                    print(_)
                    break  # 找到后退出循环
                if not found:
                    print("查询失败")

        if self.model == 'excel':
            list_data = self.tuple_to_dict(self.loading_data_list)
            for _ in list_data:
                if self.find_name == _["姓名"]:
                    print("成功查询")
                    found = True
                    print(_)
                    break
                if not found:
                    print("查询失败")

    def find_by_id(self):
        found = False
        if self.model == 'txt':
            for _ in self.loading_data_list:
                dict_data = self.list_to_dict(_)
                if self.find_id == dict_data["学号"]:
                    print("成功查询")
                    found = True
                    print(_)
                    break
            if not found:
                print("查询失败")
        if self.model == 'excel':
            list_data = self.tuple_to_dict(self.loading_data_list)
            for _ in list_data:
                if self.find_id == _["学号"]:
                    print("成功查询")
                    found = True
                    print(str(_).strip('{}'))
                    break
            if not found:
                print("查询失败")


    """
        补充删除处理excel文件功能
    """
    def drop_information(self):
        flag = False
        while not flag:
            _ = Menu().menu_del()
            drop_op = input("请输入你的操作:")
            if drop_op == '1':
                self.del_id = input("请输入你要删除的学号:")
                found = False
                self.model = self.knew_txt_or_excel()
                if self.model == None:
                    flag = True
                if self.model == 'txt':
                    for _ in self.loading_data_list:
                        dict_data = self.list_to_dict(_)
                        if self.del_id == dict_data["学号"]:
                            print("成功锁定")
                            found = True
                            print(_)
                            del_sure = input("你确实要删除吗?[y/n]")
                            if del_sure != 'y':
                                flag = True
                            else:
                                self.drop_open_txt()
                                flag = True
                #  添加处理excel功能
                if self.model == 'excel':
                    data = self.tuple_to_dict(self.loading_data_list)
                    index = 0
                    for item in data:
                        index += 1
                        if self.del_id == item["学号"]:
                            print("成功锁定")
                            found = True
                            print(item)
                            del_sure = input("你确实要删除吗?[y/n]")
                            if del_sure != 'y':
                                flag = True
                            else:
                                self.drop_open_excel(index)
                                flag = True
                if not found:
                    print("删除失败")

            if drop_op == '2':
                break

    def drop_open_excel(self,index):
        wb = load_workbook('./学生数据文件/学生数据文件.xlsx')
        ws = wb.active
        ws.delete_rows(index)
        for row in range(2, ws.max_row + 1):
            ws.cell(row=row, column=1, value=row)
        # 保存对 Excel 文件的修改
        wb.save('./学生数据文件/学生数据文件.xlsx')

    def drop_open_txt(self):
        with open(f'./学生数据文件/{self.file_name_txt}','r',encoding='utf-8') as f:
            lines = f.readlines()
        with open(f'./学生数据文件/{self.file_name_txt}','w',encoding='utf-8') as f:
            for line in lines:
                # 转字典形式
                if self.del_id in line:
                    pass
                else:
                    f.write(line)
